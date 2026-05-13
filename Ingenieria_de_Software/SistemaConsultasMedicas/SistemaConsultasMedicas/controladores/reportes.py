"""
controladores/reportes.py - Generación real de PDF y Excel
=========================================================
Usa reportlab y openpyxl. Si no están instalados, lanza un mensaje claro.
"""
from datetime import datetime


def exportar_pdf(ruta, titulo, encabezados, filas, subtitulo=""):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
    except ImportError as e:
        raise RuntimeError(
            "Falta instalar reportlab. Ejecuta:  pip install reportlab"
        ) from e

    doc = SimpleDocTemplate(ruta, pagesize=landscape(letter),
                            leftMargin=30, rightMargin=30,
                            topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"<b>{titulo}</b>", styles["Title"]))
    if subtitulo:
        elementos.append(Paragraph(subtitulo, styles["Italic"]))
    elementos.append(Paragraph(
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"]))
    elementos.append(Spacer(1, 12))

    data = [encabezados] + [[str(c) if c is not None else "" for c in fila]
                            for fila in filas]
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E4D64")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.whitesmoke, colors.lightgrey]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(t)
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(
        "Sistema de Gestión de Consultas Médicas — "
        "Braulio Yael Carranza Zamora", styles["Italic"]))
    doc.build(elementos)


def exportar_excel(ruta, titulo, encabezados, filas):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError as e:
        raise RuntimeError(
            "Falta instalar openpyxl. Ejecuta:  pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:30]

    # Título
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(encabezados))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="0E4D64")
    c.alignment = Alignment(horizontal="center")

    # Encabezados
    for j, txt in enumerate(encabezados, 1):
        cell = ws.cell(row=3, column=j, value=txt)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0E9B83")
        cell.alignment = Alignment(horizontal="center")

    # Datos
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))
    for i, fila in enumerate(filas, 4):
        for j, val in enumerate(fila, 1):
            cell = ws.cell(row=i, column=j,
                           value="" if val is None else str(val))
            cell.border = border

    # Auto-ancho aproximado
    for col_idx, txt in enumerate(encabezados, 1):
        max_len = max([len(str(txt))] +
                      [len(str(f[col_idx - 1]) if f[col_idx - 1] is not None else "")
                       for f in filas])
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = \
            min(max(max_len + 2, 10), 40)

    wb.save(ruta)
